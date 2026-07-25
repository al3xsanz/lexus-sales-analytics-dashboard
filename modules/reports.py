from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.analysis import get_executive_summary

OUTPUT_FOLDER = Path("output")


def export_results(
    monthly_sales,
    model_sales,
    salesperson_sales,
    weekday_sales
):
    """
    Export the main analysis tables as CSV files.
    """

    OUTPUT_FOLDER.mkdir(exist_ok=True)

    monthly_sales.to_csv(
        OUTPUT_FOLDER / "monthly_sales.csv",
        index=False
    )

    model_sales.to_csv(
        OUTPUT_FOLDER / "model_sales.csv"
    )

    salesperson_sales.to_csv(
        OUTPUT_FOLDER / "salesperson_sales.csv"
    )

    weekday_sales.to_csv(
        OUTPUT_FOLDER / "weekday_sales.csv"
    )

    print("\nCSV reports exported successfully.")
    print(f"Saved inside: {OUTPUT_FOLDER}")


def export_excel_report(
    dataframe,
    monthly_sales,
    model_sales,
    salesperson_sales,
    weekday_sales
):
    """
    Create a formatted Excel management report.
    """

    OUTPUT_FOLDER.mkdir(exist_ok=True)

    report_path = (
        OUTPUT_FOLDER /
        "Lexus_Sales_Report.xlsx"
    )

    summary = get_executive_summary(
        dataframe,
        monthly_sales,
        model_sales,
        salesperson_sales,
        weekday_sales
    )

    executive_summary = pd.DataFrame(
        {
            "Metric": [
                "Total Units Sold",
                "Total Revenue",
                "Average Revenue per Vehicle",
                "Best Month by Units",
                "Best Month by Revenue",
                "Best-Selling Model",
                "Top Salesperson by Revenue",
                "Best Weekday by Revenue",
            ],
            "Value": [
                summary["total_units"],
                summary["total_revenue"],
                summary["average_revenue"],
                (
                    f"{summary['best_month_by_units']} "
                    f"({summary['best_month_units']} units)"
                ),
                (
                    f"{summary['best_month_by_revenue']} "
                    f"(${summary['best_month_revenue']:,.2f})"
                ),
                (
                    f"{summary['top_model']} "
                    f"({summary['top_model_units']} units)"
                ),
                (
                    f"{summary['top_salesperson']} "
                    f"(${summary['top_salesperson_revenue']:,.2f})"
                ),
                (
                    f"{summary['best_weekday']} "
                    f"(${summary['best_weekday_revenue']:,.2f})"
                ),
            ],
        }
    )

    with pd.ExcelWriter(
        report_path,
        engine="openpyxl"
    ) as writer:

        executive_summary.to_excel(
            writer,
            sheet_name="Executive Summary",
            index=False
        )

        monthly_sales.to_excel(
            writer,
            sheet_name="Monthly Sales",
            index=False
        )

        model_sales.reset_index().to_excel(
            writer,
            sheet_name="Model Performance",
            index=False
        )

        salesperson_sales.reset_index().to_excel(
            writer,
            sheet_name="Salespeople",
            index=False
        )

        weekday_sales.reset_index().to_excel(
            writer,
            sheet_name="Weekdays",
            index=False
        )

        dataframe.to_excel(
            writer,
            sheet_name="Combined Sales Data",
            index=False
        )

    format_excel_report(
        report_path
    )

    print("\nExcel report generated successfully.")
    print(f"Saved to: {report_path}")


def format_excel_report(report_path):
    """
    Apply formatting to the generated Excel workbook.
    """

    workbook = load_workbook(
        report_path
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    alternating_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7"
    )

    currency_headers = {
        "revenue",
        "average_revenue",
    }

    unit_headers = {
        "units",
    }

    for worksheet in workbook.worksheets:

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        worksheet.row_dimensions[1].height = 24

        for row_number in range(
            2,
            worksheet.max_row + 1
        ):
            if row_number % 2 == 0:
                for cell in worksheet[
                    row_number
                ]:
                    cell.fill = alternating_fill

        for column_cells in worksheet.columns:
            maximum_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:
                if cell.value is not None:
                    maximum_length = max(
                        maximum_length,
                        len(str(cell.value))
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                maximum_length + 3,
                40
            )

        headers = {
            cell.column: (
                str(cell.value)
                .strip()
                .lower()
            )
            for cell in worksheet[1]
            if cell.value is not None
        }

        for column_number, header in headers.items():

            if header in currency_headers:
                for row_number in range(
                    2,
                    worksheet.max_row + 1
                ):
                    worksheet.cell(
                        row=row_number,
                        column=column_number
                    ).number_format = (
                        '$#,##0.00'
                    )

            if header in unit_headers:
                for row_number in range(
                    2,
                    worksheet.max_row + 1
                ):
                    worksheet.cell(
                        row=row_number,
                        column=column_number
                    ).number_format = (
                        '#,##0'
                    )

            if "growth_pct" in header:
                for row_number in range(
                    2,
                    worksheet.max_row + 1
                ):
                    worksheet.cell(
                        row=row_number,
                        column=column_number
                    ).number_format = (
                        '0.00"%"'
                    )

    format_executive_summary(
        workbook
    )

    workbook.save(
        report_path
    )


def format_executive_summary(workbook):
    """
    Apply special formatting to the executive summary sheet.
    """

    worksheet = workbook[
        "Executive Summary"
    ]

    worksheet.column_dimensions[
        "A"
    ].width = 32

    worksheet.column_dimensions[
        "B"
    ].width = 42

    for row_number in range(
        2,
        worksheet.max_row + 1
    ):
        worksheet.cell(
            row=row_number,
            column=1
        ).font = Font(
            bold=True
        )

    # Numeric values in the summary:
    # B2 = units
    # B3 = total revenue
    # B4 = average revenue

    worksheet["B2"].number_format = (
        '#,##0'
    )

    worksheet["B3"].number_format = (
        '$#,##0.00'
    )

    worksheet["B4"].number_format = (
        '$#,##0.00'
    )

    worksheet.freeze_panes = "A2"